"""
Generates the manual QA workbook for the (now React-based) Landing Pages flow.
Run: python docs/make_landing_test_plan.py
Output: docs/Landing-Pages-Manual-Test-Plan.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ACCENT = "1693C9"
NAVY = "111315"
SOFT = "F4F5F7"
LINE = "E4E7EB"
GREENH = "1F8F3A"

hdr_fill = PatternFill("solid", fgColor=ACCENT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, color=NAVY, size=16)
sub_font = Font(color="5F656D", size=10)
sect_font = Font(bold=True, color="FFFFFF", size=11)
sect_fill = PatternFill("solid", fgColor=NAVY)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border


def section_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = sect_fill
    cell.font = sect_font
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, ncols, height=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = wrap_top
        cell.border = border
        cell.font = Font(size=10, color=NAVY)
    if height:
        ws.row_dimensions[row].height = height


wb = Workbook()

# ───────────────────────── Sheet 1: How to use ─────────────────────────
ws = wb.active
ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws["A1"] = "Landing Pages — Manual Test Plan"
ws["A1"].font = title_font
ws["A2"] = "Covers the converted, fully React-based public landing pages (classic + video-landing) and the lead flow."
ws["A2"].font = sub_font
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 96

rows = [
    ("", ""),
    ("WHAT CHANGED", "The public landing page (/l/{slug}) and the get-started flow are now rendered in REACT, not Blade. "
                     "The editor (/crm/landing-pages) was already React and is unchanged. Blade is now only used for agent websites."),
    ("Environment", "App URL: http://bunnyidx.test   |   Editor: http://bunnyidx.test/crm/landing-pages"),
    ("Run the app", "Make sure the dev stack is running: `composer dev` (server + queue + vite + reverb). "
                    "Vite must be running (HMR) OR a production build done (`npm run build`)."),
    ("Public page URL", "After publishing a page, its public URL is http://bunnyidx.test/l/{slug}"),
    ("Flow URL", "The full-screen questionnaire is http://bunnyidx.test/l/{slug}/get-started"),
    ("Legend", "Pass/Fail column has a dropdown. Fill the 'Actual result / Notes' column when something differs."),
    ("Tabs", "1) How to use   2) Test Plan   3) Sample Lead Data   4) React-vs-Blade Check"),
    ("Most important", "Tab 4 tells you exactly how to confirm the page is React and NOT Blade (view-source check)."),
]
r = 3
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, color=ACCENT, size=10)
    ws.cell(row=r, column=1).alignment = wrap_top
    ws.cell(row=r, column=2, value=b).alignment = wrap_top
    ws.cell(row=r, column=2).font = Font(size=10, color=NAVY)
    if b and len(b) > 90:
        ws.row_dimensions[r].height = 46
    r += 1

# ───────────────────────── Sheet 2: Test Plan ─────────────────────────
ws = wb.create_sheet("Test Plan")
ws.sheet_view.showGridLines = False
cols = ["#", "Area", "Test case", "Steps", "Test data", "Expected result", "Pass / Fail", "Actual result / Notes"]
widths = [5, 16, 30, 50, 26, 50, 12, 34]
ws.append(cols)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
style_header(ws, 1, len(cols))
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 26

# (area, case, steps, data, expected)
plan = [
    ("SECTION", "A. Editor (already React — sanity that nothing broke)"),
    ("Editor", "Open Landing Pages list", "Go to /crm/landing-pages", "—",
     "List loads. If pages exist they show with Published/Draft + lead count. Empty state shows 'Create your first landing page' (blue button)."),
    ("Editor", "Create from template", "Click 'Create landing page' → pick a Classic preset → Create from preset", "Preset: 2% Listing",
     "Redirects to the block editor for the new page."),
    ("Editor", "Edit a block", "In editor, open the Hero block, change the Headline, save", "Headline: 'Test Headline 123'",
     "Change saves with no error (PATCH succeeds)."),
    ("Editor", "Publish", "Toggle Publish on", "—", "Page becomes Published; public URL works."),

    ("SECTION", "B. Public Classic page — RENDER (React)"),
    ("Classic", "Page loads", "Open /l/{slug} of a published Classic page", "—",
     "Page renders fully: header (if logo/brand set), hero with background image, all blocks, footer. No broken layout, no raw HTML."),
    ("Classic", "Hero content correct", "Look at the hero", "—",
     "Eyebrow, headline, subheadline match what you set in the editor. Address pill + button visible."),
    ("Classic", "Each block renders", "Scroll through the whole page", "—",
     "Logos, Content, About (with photo+stats), Steps, Value props, Testimonials, Calculator, Video, CTA — each renders styled, matching the old design."),
    ("Classic", "Accent color applied", "Check buttons / eyebrows / links", "Accent: #1693C9",
     "Accent color is used on buttons, eyebrows, focus rings — matches the page's accent."),
    ("Classic", "Draft is private", "Log out, open /l/{slug} of a DRAFT (unpublished) page", "—",
     "404 Not Found (drafts are not public)."),
    ("Classic", "Owner can preview draft", "Logged in as the owner, open the draft /l/{slug}", "—",
     "Page loads with a black 'Draft preview — not published yet' banner at top."),

    ("SECTION", "C. Hero capture flow (Seller / Valuation)"),
    ("Hero flow", "Address required", "On hero, click 'Get Started' with empty address", "(empty)",
     "Input is focused and shows a red outline; nothing navigates."),
    ("Hero flow", "Address autocomplete", "Type an address in the hero pill", "123 Brickell Ave, Miami",
     "Google Places suggestions appear (if Maps key configured); picking one fills the field."),
    ("Hero flow", "Confirm modal opens", "Enter an address, click 'Get Started'", "123 Brickell Ave, Miami",
     "A modal opens with a Google map of the address, the address text, and an 'Are you the owner?' dropdown."),
    ("Hero flow", "Owner question required", "In modal, click Continue without choosing an option", "—",
     "Dropdown shows a red border; does not continue."),
    ("Hero flow", "Continue → flow", "Choose 'Yes, I'm the owner' → Continue", "Yes, I'm the owner",
     "Navigates to /l/{slug}/get-started?address=...&owner=... (the full-screen questionnaire)."),
    ("Hero flow", "Edit address / close", "Reopen modal, click 'Edit address' and the X / press Esc", "—",
     "Modal closes; focus returns to the address input."),

    ("SECTION", "D. Buyer hero flow"),
    ("Buyer flow", "Buyer skips map modal", "On a Buyer-flow page, enter an area and click the button", "Miami, FL",
     "Goes straight to /l/{slug}/get-started?address=... (no map modal — by design for buyer flow)."),

    ("SECTION", "E. Full-screen questionnaire + lead submit (React)"),
    ("Flow", "Questions render", "On /l/{slug}/get-started, answer each step", "Use Sample Lead Data tab",
     "Progress bar advances; each question shows options; Back button works after step 1."),
    ("Flow", "Address chip", "Check top of the questionnaire", "—",
     "Shows the address you entered (or 'Searching: …' for buyer)."),
    ("Flow", "Contact form", "Reach the final step", "Name/email/phone from Sample Data",
     "Name, Email, Phone fields + SMS consent checkbox + submit button (label from editor)."),
    ("Flow", "Consent required", "Try to submit without checking consent", "—",
     "Browser blocks submit (consent + required fields enforced)."),
    ("Flow", "Submit creates lead", "Fill everything, check consent, submit", "See Sample Data row 1",
     "Redirects to a full-screen 'Thank you!' page with a Back-to-site button."),

    ("SECTION", "F. Inline lead form block"),
    ("Lead form", "Submit inline form", "On a page with a Lead Capture Form block, fill + submit", "See Sample Data row 2",
     "Page reloads with a green 'Thanks! Your details were received' banner; the form area shows the thank-you state."),

    ("SECTION", "G. Calculator block (interactive)"),
    ("Calculator", "Slider recalculates", "Drag the home-value slider", "Slide to $750,000",
     "Home value, traditional commission, our commission and 'you keep' savings update live and correctly."),

    ("SECTION", "H. Video Landing template (React, Tailwind)"),
    ("Video LP", "Page renders", "Open a published Video-Landing page", "—",
     "Dark navy theme, orange accent, urgency bar, video hero, benefits grid, testimonials, guarantee, authority, application form, footer."),
    ("Video LP", "Video plays", "Click the video / play button", "—",
     "Video opens / plays (YouTube/Vimeo embed) if a video URL is set; else poster shows."),
    ("Video LP", "Anchor buttons", "Click an 'Apply' CTA in hero/guarantee/authority", "—",
     "Page scrolls to the application form (#apply)."),
    ("Video LP", "Application submit", "Fill + submit the application form", "See Sample Data row 3",
     "Reloads with the green success banner / thank-you state; lead is created."),

    ("SECTION", "I. Lead reaches the CRM"),
    ("CRM", "Contact created", "After any submit above, go to /crm (Contacts)", "—",
     "A new Contact exists with the right name/email/phone/address, Source = 'Landing Page', correct type (buyer/seller)."),
    ("CRM", "Questionnaire saved", "Open that contact", "—",
     "Description contains the questionnaire answers + property address."),
    ("CRM", "UTM captured", "Open a /l/{slug}?utm_source=google&utm_campaign=test page, submit a lead, open the contact", "?utm_source=google&utm_campaign=test",
     "Contact custom fields show utm_source=google, utm_campaign=test."),
    ("CRM", "Lead count", "Reload the Landing Pages list", "—",
     "The page's lead count increased by the number of test submissions."),

    ("SECTION", "J. SEO / meta (server-rendered in the React shell)"),
    ("SEO", "Title + description", "Open /l/{slug}, view page source (Ctrl+U)", "—",
     "<title> and <meta name=description> in the HTML match the page's SEO settings (server-rendered, not blank)."),
    ("SEO", "Draft noindex", "View source of a draft preview", "—",
     "<meta name=robots content='noindex, nofollow'> is present for drafts and for the flow page."),
]

r = 2
n = 1
for item in plan:
    if item[0] == "SECTION":
        section_row(ws, r, item[1], len(cols))
        r += 1
        continue
    area, case, steps, data, expected = item
    ws.cell(row=r, column=1, value=n)
    ws.cell(row=r, column=2, value=area)
    ws.cell(row=r, column=3, value=case)
    ws.cell(row=r, column=4, value=steps)
    ws.cell(row=r, column=5, value=data)
    ws.cell(row=r, column=6, value=expected)
    ws.cell(row=r, column=7, value="")
    ws.cell(row=r, column=8, value="")
    est = max(len(steps), len(expected))
    body(ws, r, len(cols), height=30 if est < 55 else (46 if est < 95 else 62))
    ws.cell(row=r, column=1).alignment = center
    ws.cell(row=r, column=7).alignment = center
    n += 1
    r += 1

# Pass/Fail dropdown
dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"G2:G{r-1}")

# ───────────────────────── Sheet 3: Sample Lead Data ─────────────────────────
ws = wb.create_sheet("Sample Lead Data")
ws.sheet_view.showGridLines = False
ws["A1"] = "Sample lead data — paste / type these while testing the forms"
ws["A1"].font = Font(bold=True, color=NAVY, size=13)
hdr = ["Row", "Full name", "Email", "Phone", "Address / Area", "Lead type", "Use for"]
ws.append([])
ws.append(hdr)
style_header(ws, 3, len(hdr))
data_rows = [
    [1, "John Seller", "john.seller@example.com", "305-555-1001", "123 Brickell Ave, Miami, FL", "seller", "Seller hero → questionnaire → submit"],
    [2, "Maria Buyer", "maria.buyer@example.com", "305-555-1002", "Coral Gables, FL", "buyer", "Buyer flow + inline lead form"],
    [3, "David Apply", "david.apply@example.com", "786-555-1003", "Coconut Grove, FL", "seller", "Video-landing application form"],
    [4, "Aisha Valuation", "aisha.val@example.com", "954-555-1004", "456 Ocean Dr, Miami Beach, FL", "seller", "Home valuation flow"],
    [5, "Tom Lead", "tom.lead@example.com", "+1 305 555 1005", "789 Sunset Blvd, Hollywood, FL", "buyer", "UTM test: append ?utm_source=google&utm_campaign=test"],
]
widths3 = [6, 18, 30, 18, 32, 12, 40]
for i, w in enumerate(widths3, 1):
    ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w
rr = 4
for row in data_rows:
    ws.append(row)
    body(ws, rr, len(hdr), height=20)
    ws.cell(row=rr, column=1).alignment = center
    rr += 1
note = ws.cell(row=rr + 1, column=1,
               value="Note: emails use @example.com so they are obviously test leads. Delete these contacts from the CRM after testing.")
note.font = Font(italic=True, color="5F656D", size=10)
ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=7)

# ───────────────────────── Sheet 4: React-vs-Blade Check ─────────────────────────
ws = wb.create_sheet("React-vs-Blade Check")
ws.sheet_view.showGridLines = False
ws["A1"] = "How to confirm the landing page is REACT (not Blade)"
ws["A1"].font = title_font
ws["A2"] = "Do these checks on a published public page: http://bunnyidx.test/l/{slug}"
ws["A2"].font = sub_font
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 62
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 12

checks_hdr = ["#", "Check", "How to do it", "Expected (React) result", "Pass/Fail"]
ws.append([])
ws.append([])
ws.append(checks_hdr)
style_header(ws, 4, len(checks_hdr))
ws.freeze_panes = "A5"

checks = [
    ("View page source", "Open the page, press Ctrl+U (View Source).",
     "The <body> is almost EMPTY — basically one <div id=\"lp-root\" data-page='{...big JSON...}'></div>. The block HTML (hero text, sections) is NOT in the source — it lives inside the JSON payload, and React builds the DOM in the browser."),
    ("Old Blade markup is gone", "In View Source, Ctrl+F for 'lp-hero-inner' or section block HTML.",
     "You should NOT find the rendered block HTML in the raw source (only inside the data-page JSON). With the old Blade version the full block HTML was in the source."),
    ("React root mount", "In View Source, Ctrl+F for: id=\"lp-root\"",
     "Found exactly once. This is the React mount node."),
    ("JSON payload present", "In View Source, Ctrl+F for: data-page=",
     "Found — a big JSON blob with \"mode\":\"page\", blocks, agent, config, accent, etc."),
    ("React bundle loaded", "In View Source, look at the <script type=\"module\"> tags.",
     "A module script loads resources/js/landing-pages/public/app.tsx (dev) or /build/assets/app-*.js (prod). That is the React app."),
    ("DevTools — Elements", "Open DevTools (F12) → Elements. Compare to View Source.",
     "Elements shows the FULL rendered page (all sections) but View Source was nearly empty → proof the DOM was built client-side by React."),
    ("DevTools — Components", "Install React DevTools extension → open the 'Components' tab.",
     "A React component tree is shown (Layout/VideoLayout + the block components). Blade pages have no React tree."),
    ("Disable JavaScript test", "In DevTools, disable JavaScript, then reload the page.",
     "The page is BLANK (only the empty root). This proves rendering depends on React JS. (A Blade page would still show content with JS off.)"),
    ("SEO still server-side", "View Source, check <head>.",
     "<title> and <meta description> ARE in the raw HTML (server-rendered in the shell) even though the body is React — this is the intended hybrid."),
    ("Flow page is React too", "Open /l/{slug}/get-started, View Source.",
     "Same pattern: empty body + #lp-root + data-page JSON with \"mode\":\"flow\"."),
]
rr = 5
for i, (chk, how, exp) in enumerate(checks, 1):
    ws.cell(row=rr, column=1, value=i)
    ws.cell(row=rr, column=2, value=chk)
    ws.cell(row=rr, column=3, value=how)
    ws.cell(row=rr, column=4, value=exp)
    ws.cell(row=rr, column=5, value="")
    body(ws, rr, len(checks_hdr), height=58 if len(exp) > 120 else 44)
    ws.cell(row=rr, column=1).alignment = center
    ws.cell(row=rr, column=5).alignment = center
    rr += 1
dv2 = DataValidation(type="list", formula1='"Pass,Fail"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add(f"E5:E{rr-1}")

verdict = ws.cell(row=rr + 1, column=2,
                  value="Bottom line: if View Source body is just #lp-root + JSON, and the visible page only appears once JS runs — it is React, not Blade. ✔")
verdict.font = Font(bold=True, color=GREENH, size=11)
ws.merge_cells(start_row=rr + 1, start_column=2, end_row=rr + 1, end_column=4)

wb.save("docs/Landing-Pages-Manual-Test-Plan.xlsx")
print("Wrote docs/Landing-Pages-Manual-Test-Plan.xlsx")
print("Sheets:", wb.sheetnames)
